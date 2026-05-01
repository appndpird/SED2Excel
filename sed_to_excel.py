import os, glob, re, threading
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import webbrowser

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False

# ── Palette ───────────────────────────────────────────────────────────────────
NAVY   = "#1A3A5C"
BLUE   = "#2E6DA4"
SKY    = "#5BA3D9"
LIGHT  = "#EEF4FB"
WHITE  = "#FFFFFF"
ACCENT = "#F0A500"
TEXT   = "#1C2B3A"
GRAY   = "#8A9BB0"

# ── Metadata field order (mirrors SED header) ─────────────────────────────────
METADATA_FIELDS = [
    "Comment", "Version", "File_Name", "Instrument", "Detectors",
    "Measurement", "Date", "Time", "Temperature_C", "Battery_Voltage",
    "Averages", "Integration", "Dark_Mode", "Foreoptic",
    "Radiometric_Calibration", "Units", "Wavelength_Range",
    "Latitude", "Longitude", "Altitude", "GPS_Time", "Satellites",
    "Calibrated_Reference_Correction_File", "Channels",
]

# Map from raw SED key (lowercase) to our column name
_KEY_MAP = {
    "comment":                              "Comment",
    "version":                              "Version",
    "file name":                            "File_Name",
    "instrument":                           "Instrument",
    "detectors":                            "Detectors",
    "measurement":                          "Measurement",
    "date":                                 "Date",
    "time":                                 "Time",
    "temperature (c)":                      "Temperature_C",
    "battery voltage":                      "Battery_Voltage",
    "averages":                             "Averages",
    "integration":                          "Integration",
    "dark mode":                            "Dark_Mode",
    "foreoptic":                            "Foreoptic",
    "radiometric calibration":              "Radiometric_Calibration",
    "units":                                "Units",
    "wavelength range":                     "Wavelength_Range",
    "latitude":                             "Latitude",
    "longitude":                            "Longitude",
    "altitude":                             "Altitude",
    "gps time":                             "GPS_Time",
    "satellites":                           "Satellites",
    "calibrated reference correction file": "Calibrated_Reference_Correction_File",
    "channels":                             "Channels",
}


# ══════════════════════════════════════════════════════════════════════════════
#  SED PARSER
# ══════════════════════════════════════════════════════════════════════════════
def parse_sed(filepath):
    """
    Returns (file_id, metadata, spectral) where
      file_id  : trailing numeric string from filename, e.g. '00001'
      metadata : {column_name: raw_value_string}
      spectral : {wavelength_int: float_value}
    Returns None on failure.
    """
    metadata = {}
    spectral = {}
    in_data        = False
    header_skipped = False

    try:
        with open(filepath, "r", encoding="utf-8", errors="replace") as f:
            for line in f:
                stripped = line.strip()

                if stripped.startswith("Data:"):
                    in_data = True
                    continue

                if in_data:
                    if not header_skipped:
                        header_skipped = True
                        continue
                    parts = stripped.split()
                    if len(parts) == 2:
                        try:
                            wvl = int(float(parts[0]))
                            val = float(parts[1])
                            spectral[wvl] = val
                        except ValueError:
                            pass
                else:
                    # Parse header  key: value
                    if ":" in stripped:
                        key, _, value = stripped.partition(":")
                        key_norm = key.strip().lower()
                        col_name = _KEY_MAP.get(key_norm)
                        if col_name:
                            metadata[col_name] = value.strip()

        if not spectral:
            return None

        # Extract filename stem and trailing numeric ID
        stem = os.path.splitext(os.path.basename(filepath))[0]
        m = re.search(r"(\d+)$", stem)
        numeric_id = m.group(1) if m else stem

        return numeric_id, stem, metadata, spectral

    except Exception:
        return None


# ══════════════════════════════════════════════════════════════════════════════
#  EXCEL STYLE HELPERS
# ══════════════════════════════════════════════════════════════════════════════
def _make_styles():
    thin = Side(style="thin", color="BDD7EE")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    return {
        "hdr_fill":   PatternFill("solid", fgColor="1A3A5C"),
        "id_fill":    PatternFill("solid", fgColor="2E6DA4"),
        "alt_fill":   PatternFill("solid", fgColor="EEF4FB"),
        "white_fill": PatternFill("solid", fgColor="FFFFFF"),
        "hdr_font":   Font(name="Calibri", bold=True, color="FFFFFF", size=11),
        "id_font":    Font(name="Calibri", bold=True, color="FFFFFF", size=10),
        "body_font":  Font(name="Calibri", size=10),
        "border":     border,
        "center":     Alignment(horizontal="center", vertical="center"),
        "left":       Alignment(horizontal="left",   vertical="center"),
    }


def _apply(cell, fill, font, alignment, border):
    cell.fill = fill; cell.font = font
    cell.alignment = alignment; cell.border = border


def _write_sheet(ws, col_headers, rows, s, meta_col_count=0):
    """
    col_headers    : list of strings (first = 'File_ID')
    rows           : list of lists aligned with col_headers
    meta_col_count : how many columns after File_ID are metadata (for wider cols)
    """
    for ci, h in enumerate(col_headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        _apply(c, s["hdr_fill"], s["hdr_font"], s["center"], s["border"])

    for ri, row_vals in enumerate(rows, 2):
        fill = s["alt_fill"] if ri % 2 == 0 else s["white_fill"]
        for ci, val in enumerate(row_vals, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            if ci <= 2:   # ID and File_ID both get accent style
                _apply(c, s["id_fill"], s["id_font"], s["center"], s["border"])
            else:
                _apply(c, fill, s["body_font"], s["center"], s["border"])
                if isinstance(val, float):
                    c.number_format = "0.0000"

    # Column widths: ID narrow, File_ID wider, metadata wider, spectral narrow
    ws.column_dimensions["A"].width = 10   # ID
    ws.column_dimensions["B"].width = 38   # File_ID (full filename)
    for ci, h in enumerate(col_headers[2:], 3):
        col_letter = get_column_letter(ci)
        if ci <= 2 + meta_col_count:
            ws.column_dimensions[col_letter].width = max(len(str(h)) + 4, 18)
        else:
            ws.column_dimensions[col_letter].width = 8

    ws.freeze_panes = "C2"   # freeze both ID columns
    ws.row_dimensions[1].height = 22
    for r in range(2, len(rows) + 2):
        ws.row_dimensions[r].height = 16


# ══════════════════════════════════════════════════════════════════════════════
#  EXCEL WRITERS
# ══════════════════════════════════════════════════════════════════════════════
def write_spectral_excel(records, wavelengths, out_path):
    """Spectral-only workbook: ID, File_ID, 350, 351, ..., 2500"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Spectral Data"
    s = _make_styles()

    col_headers = ["ID", "File_ID"] + [str(w) for w in wavelengths]
    rows = [[num_id, file_id] + [spec.get(w, None) for w in wavelengths]
            for num_id, file_id, _, spec in records]

    _write_sheet(ws, col_headers, rows, s, meta_col_count=0)
    wb.save(out_path)


def write_metadata_excel(records, wavelengths, out_path):
    """Full workbook: File_ID, <all SED header fields>, 350, 351, ..., 2500"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Full Data"
    s = _make_styles()

    # Only include metadata fields present in at least one record
    present_meta = [f for f in METADATA_FIELDS
                    if any(f in meta for _, _, meta, _ in records)]

    col_headers = ["ID", "File_ID"] + present_meta + [str(w) for w in wavelengths]
    rows = [
        [num_id, file_id]
        + [meta.get(f, None) for f in present_meta]
        + [spec.get(w, None) for w in wavelengths]
        for num_id, file_id, meta, spec in records
    ]

    _write_sheet(ws, col_headers, rows, s, meta_col_count=len(present_meta))
    wb.save(out_path)


# ══════════════════════════════════════════════════════════════════════════════
#  APPLICATION
# ══════════════════════════════════════════════════════════════════════════════
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("SED Hyperspectral \u2192 Excel Converter")
        self.geometry("800x710")
        self.resizable(False, False)
        self.configure(bg=NAVY)
        self._build_ui()

    def _build_ui(self):
        # ── Banner ────────────────────────────────────────────────────────────
        banner = tk.Frame(self, bg=NAVY, padx=28, pady=16)
        banner.pack(fill="x", side="top")
        tk.Label(banner, text="SED Hyperspectral \u2192 Excel Converter",
                 font=("Georgia", 17, "bold"), bg=NAVY, fg=WHITE).pack(anchor="w")
        tk.Label(banner,
                 text="Convert NaturaSpec / SpectralEvolution .SED files into structured Excel workbooks",
                 font=("Calibri", 10), bg=NAVY, fg=SKY).pack(anchor="w", pady=(2, 0))

        # ── Credit bar (packed BEFORE card to stay pinned at bottom) ──────────
        tk.Frame(self, bg=SKY, height=1).pack(fill="x", side="bottom")
        credit_bar = tk.Frame(self, bg=NAVY, padx=20, pady=10)
        credit_bar.pack(fill="x", side="bottom")

        line1 = tk.Frame(credit_bar, bg=NAVY)
        line1.pack(anchor="center")
        tk.Label(line1, text="Developed by ",
                 font=("Calibri", 9), bg=NAVY, fg=GRAY).pack(side="left")
        tk.Label(line1, text="Bipul Neupane, PhD",
                 font=("Calibri", 9, "bold"), bg=NAVY, fg=WHITE).pack(side="left")
        tk.Label(line1, text="  |  Research Scientist, DPIRD Node, APPN",
                 font=("Calibri", 9), bg=NAVY, fg=GRAY).pack(side="left")

        line2 = tk.Frame(credit_bar, bg=NAVY)
        line2.pack(anchor="center", pady=(2, 0))
        tk.Label(line2, text="Contact: ",
                 font=("Calibri", 9), bg=NAVY, fg=GRAY).pack(side="left")
        lbl_email = tk.Label(line2, text="bipul.neupane@dpird.wa.gov.au",
                             font=("Calibri", 9, "underline"), bg=NAVY,
                             fg=SKY, cursor="hand2")
        lbl_email.pack(side="left")
        lbl_email.bind("<Button-1>",
                       lambda e: webbrowser.open("mailto:bipul.neupane@dpird.wa.gov.au"))
        tk.Label(line2, text="   |   Developed with ",
                 font=("Calibri", 9), bg=NAVY, fg=GRAY).pack(side="left")
        lbl_claude = tk.Label(line2, text="claude.ai",
                              font=("Calibri", 9, "bold underline"),
                              bg=NAVY, fg=ACCENT, cursor="hand2")
        lbl_claude.pack(side="left")
        lbl_claude.bind("<Button-1>", lambda e: webbrowser.open("https://claude.ai"))

        # ── Divider ───────────────────────────────────────────────────────────
        tk.Frame(self, bg=SKY, height=2).pack(fill="x", side="top")

        # ── Main card ─────────────────────────────────────────────────────────
        card = tk.Frame(self, bg=LIGHT, padx=28, pady=20)
        card.pack(fill="both", expand=True, padx=18, pady=12, side="top")
        card.columnconfigure(0, weight=1)

        # Input folder
        self._section_label(card, "\U0001f4c2  Input Folder  (containing .SED files)", 0)
        in_fr = tk.Frame(card, bg=LIGHT)
        in_fr.grid(row=1, column=0, sticky="ew", pady=(3, 11))
        self.in_var = tk.StringVar()
        self._entry(in_fr, self.in_var).pack(
            side="left", fill="x", expand=True, ipady=7, padx=(0, 8))
        self._btn(in_fr, "Browse", self._browse_input).pack(side="left")

        # Output folder
        self._section_label(card, "\U0001f4be  Output Folder", 2)
        out_fr = tk.Frame(card, bg=LIGHT)
        out_fr.grid(row=3, column=0, sticky="ew", pady=(3, 11))
        self.out_var = tk.StringVar()
        self._entry(out_fr, self.out_var).pack(
            side="left", fill="x", expand=True, ipady=7, padx=(0, 8))
        self._btn(out_fr, "Browse", self._browse_output).pack(side="left")

        # Filename
        self._section_label(card, "\U0001f4c4  Output File Name", 4)
        fn_fr = tk.Frame(card, bg=LIGHT)
        fn_fr.grid(row=5, column=0, sticky="ew", pady=(3, 4))
        self.fn_var = tk.StringVar(value="hyperspectral_data")
        self._entry(fn_fr, self.fn_var).pack(
            side="left", fill="x", expand=True, ipady=7, padx=(0, 6))
        tk.Label(fn_fr, text=".xlsx", font=("Calibri", 10),
                 bg=LIGHT, fg=GRAY).pack(side="left")

        # Output preview box
        prev_fr = tk.Frame(card, bg="#D8EAF8", padx=12, pady=8)
        prev_fr.grid(row=6, column=0, sticky="ew", pady=(2, 14))
        tk.Label(prev_fr, text="Output files that will be created:",
                 font=("Calibri", 9, "bold"), bg="#D8EAF8", fg=NAVY).pack(anchor="w")
        self.preview1_var = tk.StringVar()
        self.preview2_var = tk.StringVar()
        tk.Label(prev_fr, textvariable=self.preview1_var,
                 font=("Courier New", 9), bg="#D8EAF8", fg=TEXT).pack(anchor="w")
        tk.Label(prev_fr, textvariable=self.preview2_var,
                 font=("Courier New", 9), bg="#D8EAF8", fg=TEXT).pack(anchor="w")
        self._update_preview()
        self.fn_var.trace_add("write", lambda *_: self._update_preview())

        # Convert button
        self.btn_convert = tk.Button(
            card, text="\u26a1  Convert SED Files to Excel",
            font=("Calibri", 12, "bold"), bg=ACCENT, fg=NAVY,
            activebackground="#D4920A", activeforeground=NAVY,
            relief="flat", cursor="hand2", padx=28, pady=10,
            command=self._run)
        self.btn_convert.grid(row=7, column=0, pady=(0, 10))

        # Progress bar
        style = ttk.Style(self)
        style.theme_use("default")
        style.configure("A.Horizontal.TProgressbar",
                        troughcolor=WHITE, background=BLUE,
                        thickness=10, borderwidth=0)
        self.progress = ttk.Progressbar(
            card, style="A.Horizontal.TProgressbar",
            orient="horizontal", mode="determinate", length=720)
        self.progress.grid(row=8, column=0, sticky="ew", pady=(0, 5))

        # Status label
        self.status_var = tk.StringVar(
            value="Ready \u2014 select an input folder to begin.")
        tk.Label(card, textvariable=self.status_var,
                 font=("Calibri", 10), bg=LIGHT, fg=TEXT,
                 wraplength=740, justify="left").grid(row=9, column=0, sticky="w")

        # Log console
        log_fr = tk.Frame(card, bg=LIGHT)
        log_fr.grid(row=10, column=0, sticky="ew", pady=(8, 0))
        self.log = tk.Text(log_fr, height=6, font=("Courier New", 9),
                           bg="#0E2235", fg="#7DD3FC", relief="flat",
                           insertbackground=WHITE, state="disabled",
                           wrap="word", padx=10, pady=8)
        sc = tk.Scrollbar(log_fr, command=self.log.yview, bg=NAVY)
        self.log.configure(yscrollcommand=sc.set)
        sc.pack(side="right", fill="y")
        self.log.pack(side="left", fill="both", expand=True)

    # ── Widget helpers ────────────────────────────────────────────────────────
    def _section_label(self, p, text, row):
        tk.Label(p, text=text, font=("Calibri", 11, "bold"),
                 bg=LIGHT, fg=NAVY).grid(row=row, column=0, sticky="w")

    def _entry(self, parent, var):
        return tk.Entry(parent, textvariable=var, font=("Calibri", 10),
                        relief="flat", bg=WHITE, fg=TEXT, insertbackground=TEXT,
                        highlightthickness=1, highlightbackground=BLUE,
                        highlightcolor=SKY)

    def _btn(self, parent, label, cmd):
        return tk.Button(parent, text=label, font=("Calibri", 10, "bold"),
                         bg=BLUE, fg=WHITE, activebackground=SKY,
                         activeforeground=WHITE, relief="flat",
                         cursor="hand2", padx=14, pady=5, command=cmd)

    def _update_preview(self):
        fn = self.fn_var.get().strip() or "hyperspectral_data"
        self.preview1_var.set(
            f"  \u2022  {fn}.xlsx  \u2192  spectral data only  (File_ID + wavelength columns)")
        self.preview2_var.set(
            f"  \u2022  metadata_{fn}.xlsx  \u2192  all SED header fields + wavelength columns")

    # ── Folder pickers ────────────────────────────────────────────────────────
    def _browse_input(self):
        d = filedialog.askdirectory(title="Select folder containing .SED files")
        if d:
            self.in_var.set(d)
            if not self.out_var.get():
                self.out_var.set(d)
            n = len(set(
                glob.glob(os.path.join(d, "**", "*.sed"), recursive=True) +
                glob.glob(os.path.join(d, "**", "*.SED"), recursive=True)))
            self._log(f"Found {n} .SED file(s) in: {d}")
            self.status_var.set(
                f"Found {n} .SED file(s). Set output folder and click Convert.")

    def _browse_output(self):
        d = filedialog.askdirectory(title="Select output folder")
        if d:
            self.out_var.set(d)

    # ── Log ───────────────────────────────────────────────────────────────────
    def _log(self, msg):
        self.log.configure(state="normal")
        self.log.insert("end", msg + "\n")
        self.log.see("end")
        self.log.configure(state="disabled")

    # ── Convert ───────────────────────────────────────────────────────────────
    def _run(self):
        if not OPENPYXL_OK:
            messagebox.showerror("Missing Dependency",
                "The 'openpyxl' package is required.\n\nInstall with:\n  pip install openpyxl")
            return
        in_dir  = self.in_var.get().strip()
        out_dir = self.out_var.get().strip()
        fname   = self.fn_var.get().strip() or "hyperspectral_data"
        if not in_dir or not os.path.isdir(in_dir):
            messagebox.showerror("Error", "Please select a valid input folder.")
            return
        if not out_dir:
            messagebox.showerror("Error", "Please select an output folder.")
            return
        os.makedirs(out_dir, exist_ok=True)

        # ── Check for existing output files and ask before overwriting ────────
        spec_path = os.path.join(out_dir, f"{fname}.xlsx")
        meta_path = os.path.join(out_dir, f"metadata_{fname}.xlsx")
        existing  = [os.path.basename(p) for p in (spec_path, meta_path)
                     if os.path.exists(p)]
        if existing:
            files_list = "\n  \u2022  ".join(existing)
            answer = messagebox.askyesno(
                "Overwrite Existing Files?",
                f"The following file(s) already exist in the output folder:\n\n"
                f"  \u2022  {files_list}\n\n"
                f"Do you want to replace them?",
                icon="warning"
            )
            if not answer:
                return   # user chose No — abort silently

        self.btn_convert.configure(state="disabled", text="Processing\u2026")
        self.progress["value"] = 0
        threading.Thread(target=self._worker,
                         args=(in_dir, out_dir, fname), daemon=True).start()

    def _worker(self, in_dir, out_dir, fname):
        try:
            raw = (glob.glob(os.path.join(in_dir, "**", "*.sed"), recursive=True) +
                   glob.glob(os.path.join(in_dir, "**", "*.SED"), recursive=True))
            seen = set(); files = []
            for f in sorted(raw):
                k = os.path.normcase(os.path.abspath(f))
                if k not in seen:
                    seen.add(k); files.append(f)

            if not files:
                self.after(0, lambda: messagebox.showwarning(
                    "No Files", "No .SED files found in the selected folder."))
                self.after(0, self._reset_btn)
                return

            self._log(f"\n{'─'*60}")
            self._log(f"Processing {len(files)} file(s)\u2026")

            records = []
            all_wvl = set()
            errors  = []

            for i, fp in enumerate(files, 1):
                result = parse_sed(fp)
                name   = os.path.basename(fp)
                if result:
                    num_id, file_id, meta, spec = result
                    records.append((num_id, file_id, meta, spec))
                    all_wvl.update(spec.keys())
                    self._log(f"  [{i:>4}/{len(files)}] \u2713  {name}  "
                              f"\u2192  ID: {num_id}  File_ID: {file_id}  ({len(spec)} bands)")
                else:
                    errors.append(name)
                    self._log(f"  [{i:>4}/{len(files)}] \u2717  {name}  \u2014 parse failed")
                self.after(0, lambda v=int(i / len(files) * 75):
                           self.progress.configure(value=v))

            if not records:
                self.after(0, lambda: messagebox.showerror(
                    "Parse Error", "Could not parse any SED files."))
                self.after(0, self._reset_btn)
                return

            wavelengths = sorted(all_wvl)
            self._log(f"\nWavelength range : {wavelengths[0]}\u2013{wavelengths[-1]} nm  "
                      f"({len(wavelengths)} bands)")
            self._log(f"Samples parsed   : {len(records)}")

            # Spectral-only Excel
            spec_path = os.path.join(out_dir, f"{fname}.xlsx")
            self._log(f"\n[1/2] Writing spectral Excel  \u2192  {spec_path}")
            write_spectral_excel(records, wavelengths, spec_path)
            self.after(0, lambda: self.progress.configure(value=87))

            # Metadata + spectral Excel
            meta_path = os.path.join(out_dir, f"metadata_{fname}.xlsx")
            self._log(f"[2/2] Writing metadata Excel  \u2192  {meta_path}")
            write_metadata_excel(records, wavelengths, meta_path)
            self.after(0, lambda: self.progress.configure(value=100))

            if errors:
                self._log(f"\n\u26a0\ufe0f  {len(errors)} file(s) skipped: {', '.join(errors)}")

            self._log(f"\n\u2705  Done!  2 Excel files saved to:\n   {out_dir}")
            self.after(0, lambda: self.status_var.set(
                f"\u2705  Success! {len(records)} file(s) \u2192 2 Excel workbooks saved."))
            self.after(0, lambda: messagebox.showinfo(
                "Conversion Complete",
                f"Successfully converted {len(records)} SED file(s).\n\n"
                f"Files saved:\n"
                f"  \u2022  {fname}.xlsx\n"
                f"       (spectral data: File_ID + wavelengths)\n\n"
                f"  \u2022  metadata_{fname}.xlsx\n"
                f"       (all SED header fields + wavelengths)\n\n"
                f"Location:\n{out_dir}"))

        except Exception as exc:
            self._log(f"\n\u274c Error: {exc}")
            self.after(0, lambda: messagebox.showerror("Error", str(exc)))
        finally:
            self.after(0, self._reset_btn)

    def _reset_btn(self):
        self.btn_convert.configure(state="normal",
                                   text="\u26a1  Convert SED Files to Excel")


if __name__ == "__main__":
    App().mainloop()